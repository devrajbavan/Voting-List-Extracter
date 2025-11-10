import shutil
import os
import re
import time
from pathlib import Path
from PIL import Image, ImageEnhance
import pytesseract
from openpyxl import Workbook
from openpyxl.styles import Font
from concurrent.futures import ProcessPoolExecutor
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

# CONFIG / CONSTANTS
MARATHI_FONT = Font(name="Mangal", size=11)
MARATHI_BOLD = Font(name="Mangal", size=11, bold=True)

TESSERACT_CMD = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
pytesseract.pytesseract.tesseract_cmd = TESSERACT_CMD

IMG_SOURCE   = "voters.jpg"
IMG_DIR      = Path("temp")
EXCEL_OUT    = "voter_data.xlsx"
START_SERIAL = 9

TESS_LANG = "mar+eng"

DEFAULT_COLS = 3
DEFAULT_ROWS = 10

THUMB_W, THUMB_H = 80, 90

DEV_DIGITS = "à¥¦à¥§à¥¨à¥©à¥ªà¥«à¥¬à¥­à¥®à¥¯"
ENG_DIGITS = "0123456789"

# IMAGE CROP / PREPROCESS
def crop_all_cards():
    IMG_DIR.mkdir(exist_ok=True)
    existing = list(IMG_DIR.glob("card_*.png"))
    if existing:
        print(f"'{IMG_DIR}' already contains cropped cards â€“ skipping crop ({len(existing)} found).")
        return sorted(existing)

    big = Image.open(IMG_SOURCE)
    w, h = big.size
    cols, rows = DEFAULT_COLS, DEFAULT_ROWS
    cw, rh = w // cols, h // rows
    saved = []
    for r in range(rows):
        for c in range(cols):
            box = (c * cw, r * rh, (c + 1) * cw, (r + 1) * rh)
            card = big.crop(box)
            fn = IMG_DIR / f"card_{r+1:02d}_{c+1:02d}.png"
            card.save(fn, "PNG")
            saved.append(fn)
    print(f"Cropped {len(saved)} cards into '{IMG_DIR}'.")
    return sorted(saved)

# Preprocessing
def preprocess(img_path: Path) -> Image.Image:
    img = Image.open(img_path).convert("L")
    if img.width < 350:
        img = img.resize((img.width * 2, img.height * 2), Image.LANCZOS)
    img = ImageEnhance.Contrast(img).enhance(1.5)
    img = ImageEnhance.Sharpness(img).enhance(1.2)
    return img

# Performs OCR on a Path to a card image and returns extracted text.
def ocr_card(img_path: Path) -> str:
    try:
        img = preprocess(img_path)
        txt = pytesseract.image_to_string(img, lang=TESS_LANG, config="--psm 6")
        return txt
    except Exception as e:
        print(f"âš ï¸ OCR failed for {img_path}: {e}")
        return ""

# CLEANERS & PARSERS
def clean_voter_name(raw: str) -> str:
    s = re.sub(r'[|Â¦\\\/<>]', ' ', raw)
    s = re.sub(r'\s+[A-Za-z]{1,3}\s*', ' ', s)
    s = re.sub(r'[=z&*]', '', s)
    s = re.sub(r'\s+', ' ', s).strip()
    parts = s.split()
    if len(parts) > 4:
        s = ' '.join(parts[:4])
    return s

def clean_relative_name(raw: str) -> str:
    s = re.sub(r'[|Â¦\\\/<>]', ' ', raw)
    s = re.sub(r'\s+[A-Za-z]{1,3}\s*', ' ', s)
    s = re.sub(r'[=z&*]', '', s)
    s = re.sub(r'\s+', ' ', s).strip()
    parts = s.split()
    if len(parts) > 3:
        s = ' '.join(parts[:3])
    return s

def clean_house(text: str) -> str:
    text = re.sub(r'[^\dà¥¦-à¥¯NAna/]', '', text)
    text = text.replace('NA', 'NA').strip()
    if not text or text == 'NA':
        return 'NA'
    m = re.search(r'[\dà¥¦-à¥¯]+', text)
    return m.group(0) if m else 'NA'

def clean_age(text: str) -> str:
    m = re.search(r'[\dà¥¦-à¥¯]+', text)
    if not m:
        return ""
    age = m.group(0)
    age = "".join(ENG_DIGITS[DEV_DIGITS.index(d)] if d in DEV_DIGITS else d for d in age)
    return age

def clean_gender(code: str) -> tuple:
    code = (code or "").strip().lower()
    if 'à¤œà¥€' in code or 'à¤¸à¥à¤¤à¥à¤°à¥€' in code or 'à¤¸à¥à¤°à¥€' in code:
        return "à¤¸à¥à¤¤à¥à¤°à¥€", "à¤®à¤¹à¤¿à¤²à¤¾"
    return "à¤ªà¥", "à¤ªà¥à¤°à¥à¤·"

def parse_card(text: str) -> dict:
    data = {
        "CardID": "", "RegNo": "", "ID": "",
        "VoterName": "", "RelationLabel": "", "RelationName": "",
        "House": "NA", "Age": "", "GenderCode": "", "GenderFull": ""
    }
    if not text:
        return data

    lines = [l.strip() for l in text.split('\n') if l.strip()]

    # ID (common on first line)
    if lines:
        top_block = " ".join(lines[:3])
        m_card = re.search(r'\b[A-Z]{2,3}\d{5,10}\b', top_block)
        m_reg  = re.search(r'\b\d{1,3}/\d{1,3}/\d{1,3}\b', top_block)

        if m_card:
            data["CardID"] = m_card.group(0)
        else:
            data["CardID"] = ""

        if m_reg:
            data["RegNo"] = m_reg.group(0)
        else:
            data["RegNo"] = ""

        # remove top line after reading ID info to avoid name confusion
        lines = lines[1:]

    # Voter name (preferred pattern)
    voter_pat = re.compile(r"(?:à¤®à¤¤à¤¦à¤¾à¤°(?:à¤¾à¤šà¥‡)?\s*(?:à¤ªà¥‚à¤°à¥à¤£\s*)?)\s*[:à¤ƒ]?\s*(.+)", re.I)
    for i, line in enumerate(lines):
        m = voter_pat.search(line)
        if m:
            data["VoterName"] = clean_voter_name(m.group(1))
            del lines[i]
            break
    else:
        # fallback: any 'à¤¨à¤¾à¤µ' line not containing house/gender/age
        for i, line in enumerate(lines):
            if re.search(r"(à¤ªà¤¤à¥€|à¤µà¤¡à¤¿à¤²|à¤†à¤ˆ|à¤¸à¤¾à¤¸à¥‚|à¤ªà¤¤à¥à¤¨à¥€|à¤¸à¥‚à¤¨)", line):
                continue
            if 'à¤¨à¤¾à¤µ' in line and not re.search(r"(à¤˜à¤°|à¤²à¤¿à¤‚à¤—|à¤µà¤¯)", line):
                parts = re.split(r'[:à¤ƒ]', line, maxsplit=1)
                candidate = parts[1] if len(parts) > 1 else re.split(r'à¤¨à¤¾à¤µ', line, maxsplit=1)[-1]
                data["VoterName"] = clean_voter_name(candidate)
                del lines[i]
                break

    # Relative
    rel_pat_pati   = re.compile(r"à¤ªà¤¤à¥€à¤šà¥‡\s*à¤¨à¤¾à¤µ\s*[:à¤ƒ]?\s*(.+)", re.I)
    rel_pat_vadil  = re.compile(r"à¤µà¤¡à¤¿à¤²(?:à¤¾à¤‚à¤šà¥‡|à¥‡)\s*à¤¨à¤¾à¤µ\s*[:à¤ƒ]?\s*(.+)", re.I)
    for i, line in enumerate(lines):
        m_p = rel_pat_pati.search(line)
        if m_p:
            data["RelationLabel"] = "à¤ªà¤¤à¥€à¤šà¥‡ à¤¨à¤¾à¤µ"
            data["RelationName"] = clean_relative_name(m_p.group(1))
            del lines[i]
            break
        m_v = rel_pat_vadil.search(line)
        if m_v:
            data["RelationLabel"] = "à¤µà¤¡à¤¿à¤²à¤¾à¤‚à¤šà¥‡ à¤¨à¤¾à¤µ"
            data["RelationName"] = clean_relative_name(m_v.group(1))
            del lines[i]
            break

    # House
    m = re.search(r"à¤˜à¤°\s*à¤•à¥à¤°à¤®à¤¾à¤‚à¤•\s*[:à¤ƒ]?\s*(.+)", text, re.I)
    if m:
        data["House"] = clean_house(m.group(1))

    # Age
    m = re.search(r"à¤µà¤¯\s*[:à¤ƒ]?\s*([\dà¥¦-à¥¯]+)", text, re.I)
    if m:
        data["Age"] = clean_age(m.group(1))

    # Gender
    m = re.search(r"à¤²à¤¿à¤‚à¤—\s*[:à¤ƒ]?\s*([^\s\r\n]+)", text, re.I)
    if m:
        code_raw = m.group(1)
        data["GenderCode"], data["GenderFull"] = clean_gender(code_raw)
    else:
        data["GenderCode"], data["GenderFull"] = "à¤ªà¥", "à¤ªà¥à¤°à¥à¤·"

    return data

# FACE CROP
def crop_person_photo(card_path: Path) -> Path:
    img = Image.open(card_path)
    w, h = img.size

    left   = int(w * 0.78)
    top    = int(h * 0.30)
    right  = int(w * 0.98)
    bottom = int(h * 0.85)

    left, top = max(0, left), max(0, top)
    right, bottom = min(w, right), min(h, bottom)
    if right <= left or bottom <= top:
        return None

    face = img.crop((left, top, right, bottom))
    face = ImageEnhance.Contrast(face).enhance(1.2)
    face = ImageEnhance.Sharpness(face).enhance(1.3)

    out_path = card_path.parent / f"face_{card_path.stem}.png"
    face.save(out_path, "PNG")
    return out_path

# EXCEL GENERATION
def generate_excel_from_cards(start_serial: int = START_SERIAL):
    card_files = crop_all_cards()

    if not card_files:
        raise FileNotFoundError("No card images found and source crop failed.")

    cpu_count = max(1, os.cpu_count() or 1)
    with ProcessPoolExecutor(max_workers=cpu_count) as executor:
        ocr_results = list(executor.map(ocr_card, card_files))

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    headers = [
        "S.No.", "Card ID", "Reg. No.", "à¤®à¤¤à¤¦à¤¾à¤°à¤¾à¤šà¥‡ à¤ªà¥‚à¤°à¥à¤£: à¤¨à¤¾à¤µ",
        "à¤ªà¤¤à¥€à¤šà¥‡ à¤¨à¤¾à¤µ / à¤µà¤¡à¤¿à¤²à¤¾à¤‚à¤šà¥‡ à¤¨à¤¾à¤µ", "à¤˜à¤° à¤•à¥à¤°à¤®à¤¾à¤‚à¤• :", "à¤µà¤¯ :", "à¤²à¤¿à¤‚à¤— :", "Face image"
    ]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(1, col_idx, header)
        cell.font = MARATHI_BOLD

    # 4) Fill rows
    row = 2
    serial = start_serial
    processed = 0

    for idx, card_path in enumerate(card_files):
        txt = ocr_results[idx] if idx < len(ocr_results) else ""
        d = parse_card(txt)

        # populate text columns
        ws.cell(row, 1, idx + 1).font = MARATHI_FONT
        ws.cell(row, 2, d.get("CardID", "")).font = MARATHI_FONT
        ws.cell(row, 3, d.get("RegNo", "")).font = MARATHI_FONT
        ws.cell(row, 4, d.get("VoterName", "")).font = MARATHI_FONT
        ws.cell(row, 5, d.get("RelationName", "")).font = MARATHI_FONT
        ws.cell(row, 6, d.get("House", "")).font = MARATHI_FONT
        ws.cell(row, 7, d.get("Age", "")).font = MARATHI_FONT
        ws.cell(row, 8, d.get("GenderFull", "à¤ªà¥à¤°à¥à¤·")).font = MARATHI_FONT

        # crop face and embed image (if crop produced a file)
        face_path = None
        try:
            face_path = crop_person_photo(card_path)
        except Exception as e:
            face_path = None
            print(f"âš ï¸ Face crop failed for {card_path}: {e}")

        if face_path and face_path.exists():
            try:
                img_for_excel = XLImage(str(face_path))
                img_for_excel.width, img_for_excel.height = THUMB_W, THUMB_H
                cell_ref = f"I{row}"
                ws.add_image(img_for_excel, cell_ref)
                ws.row_dimensions[row].height = img_for_excel.height * 0.75
            except Exception as e:
                print(f"âš ï¸ Adding image failed for row {row}: {e}")

        row += 1
        serial += 1
        processed += 1

    # 5) Adjust column widths (image column I wider)
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        if col_letter == 'I':
            ws.column_dimensions['I'].width = max(18, int(THUMB_W * 0.14) + 2)
            continue
        max_len = max((len(str(c.value)) for c in col if c.value), default=0)
        ws.column_dimensions[col_letter].width = min(max_len + 3, 60)

    wb.save(EXCEL_OUT)
    print(f"\nExcel saved: {EXCEL_OUT}")
    print(f"Processed: {processed} voters. First serial: {START_SERIAL}")

    # Cleanup temp images
    cleanup_images()

# CLEANUP
def cleanup_images():
    try:
        if IMG_DIR.exists():
            time.sleep(0.1)
            shutil.rmtree(IMG_DIR)
            print(f"ðŸ§¹ Cleaned up temporary folder: {IMG_DIR}")
    except Exception as e:
        print(f"âš ï¸ Cleanup failed: {e}")

# RUN

if __name__ == "__main__":
    if not Path(IMG_SOURCE).exists():
        print(f"ERROR: Source image '{IMG_SOURCE}' not found. Place voters.jpg in script folder.")
    else:
        generate_excel_from_cards()
