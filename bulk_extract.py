import os
import re
import shutil
import time
from pathlib import Path
from concurrent.futures import ProcessPoolExecutor

from PIL import Image, ImageEnhance
import pytesseract

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage


# =========================================================
# CONFIG
# =========================================================

TESSERACT_CMD = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
pytesseract.pytesseract.tesseract_cmd = TESSERACT_CMD
TESS_LANG = "mar+eng"

INPUT_IMAGE_FOLDER = "Images/"

PAGE_DIR = Path("output_voters/cropped_pages")
CARD_DIR = Path("output_voters")
EXCEL_OUT = "voter_data_bulk.xlsx"

CROP_TOP = 340
CROP_LEFT = 60
CROP_RIGHT = 88
CROP_BOTTOM = 3270

COLS = 3
ROWS = 10

THUMB_W, THUMB_H = 80, 90

MARATHI_FONT = Font(name="Mangal", size=11)
MARATHI_BOLD = Font(name="Mangal", size=11, bold=True)

DEV_DIGITS = "०१२३४५६७८९"
ENG_DIGITS = "0123456789"


# =========================================================
# OCR PREPROCESS
# =========================================================

def preprocess(img_path: Path) -> Image.Image:
    img = Image.open(img_path).convert("L")
    if img.width < 350:
        img = img.resize((img.width * 2, img.height * 2), Image.LANCZOS)
    img = ImageEnhance.Contrast(img).enhance(1.4)
    img = ImageEnhance.Sharpness(img).enhance(1.1)
    return img


def ocr_card(img_path: Path) -> str:
    try:
        img = preprocess(img_path)
        return pytesseract.image_to_string(img, lang=TESS_LANG, config="--psm 6")
    except Exception as e:
        print(f"OCR FAILED: {img_path} → {e}")
        return ""


# =========================================================
# PAGE CROPPER
# =========================================================

def crop_page(img_path, output_folder):
    os.makedirs(output_folder, exist_ok=True)
    img = Image.open(img_path)

    w, h = img.size
    left, top = CROP_LEFT, CROP_TOP
    right, bottom = w - CROP_RIGHT, CROP_BOTTOM

    cropped = img.crop((left, top, right, bottom))

    out_path = output_folder / f"{Path(img_path).stem}_cropped.jpg"
    cropped.save(out_path, "JPEG", quality=95)

    return out_path


# =========================================================
# SPLIT INTO 30 CARDS (3x10)
# =========================================================

def split_page_into_cards(page_path: Path, output_folder: Path, cols=3, rows=10):
    img = Image.open(page_path)
    w, h = img.size
    cw, rh = w // cols, h // rows

    os.makedirs(output_folder, exist_ok=True)

    saved = []
    for r in range(rows):
        for c in range(cols):
            box = (c * cw, r * rh, (c + 1) * cw, (r + 1) * rh)
            card = img.crop(box)
            fn = output_folder / f"{page_path.stem}_card_{r+1:02d}_{c+1:02d}.png"
            card.save(fn, "PNG")
            saved.append(fn)

    return saved


# =========================================================
# NORMALIZATION HELPERS
# =========================================================

def normalize_digits(s):
    """Convert Marathi digits → English digits."""
    return "".join(
        ENG_DIGITS[DEV_DIGITS.index(ch)] if ch in DEV_DIGITS else ch
        for ch in s
    )


def normalize_ocr_id(s):
    """Fix common OCR mistakes for CardID and RegNo."""
    if not s:
        return ""

    s = s.upper()

    s = s.replace("O", "0")
    s = s.replace("I", "1")
    s = s.replace("L", "1")
    s = s.replace("B", "8")
    s = s.replace("S", "5")
    s = s.replace("G", "6")

    return re.sub(r"[^A-Z0-9/]", "", s)


def clean_voter_name(raw: str) -> str:
    raw = re.sub(r'[|¦\\\/<>]', ' ', raw)
    raw = re.sub(r'\s+', ' ', raw).strip()
    parts = raw.split()
    return ' '.join(parts[:4])


def clean_relative_name(raw: str) -> str:
    raw = re.sub(r'[|¦\\\/<>]', ' ', raw)
    raw = re.sub(r'\s+', ' ', raw).strip()
    parts = raw.split()
    return ' '.join(parts[:3])


def clean_house(t: str) -> str:
    if not t:
        return "NA"
    t = normalize_digits(t)
    m = re.search(r"[0-9]+", t)
    return m.group(0) if m else "NA"


def clean_age(t: str) -> str:
    if not t:
        return ""
    t = normalize_digits(t)
    m = re.search(r"[0-9]+", t)
    return m.group(0) if m else ""


# =========================================================
# OCR-PROOF ID + REGNO EXTRACTION
# =========================================================

def extract_card_id(text):
    """
    Very strong CardID extractor:
    Pattern: 2–4 letters + 6–10 alphanumeric.
    """
    text_n = normalize_ocr_id(text)

    m = re.findall(r"[A-Z]{2,4}[0-9A-Z]{6,10}", text_n)
    if not m:
        return ""

    return normalize_ocr_id(m[0])


def extract_reg_no(text):
    """
    Extract RegNo patterns like 113/236/1277 with OCR noise tolerance.
    """
    text_n = normalize_ocr_id(text)

    m = re.findall(r"[0-9]{1,3}/[0-9]{1,3}/[0-9]{1,5}", text_n)
    if not m:
        return ""

    return m[0]


# =========================================================
# ADVANCED PARSER
# =========================================================

def normalize_marathi(text):
    if not text:
        return ""
    rep = {
        "मतदार": ["मतरार", "मतदर", "मतदाराचे", "मतदाराचे पूर्ण"],
        "पती": ["पति", "पंत", "पवी"],
        "वडिल": ["वडील", "वडिळ", "वडी"],
        "आई": ["अई", "अइ", "ऐ"],
        "घर क्रमांक": ["घर क्र", "घर न", "घर क्र.", "घर क्रम"],
        "वय": ["य", "वय:", "वम"],
        "लिंग": ["लिग", "लिं", "लखग"],
    }
    for correct, wrongs in rep.items():
        for w in wrongs:
            text = text.replace(w, correct)
    return text


def extract_after_keyword(lines, keys):
    for line in lines:
        n = normalize_marathi(line)
        for k in keys:
            if k in n:
                return n.split(k)[-1].strip(" :")
    return ""


def parse_card(text: str) -> dict:
    data = {
        "CardID": "",
        "RegNo": "",
        "VoterName": "",
        "RelationName": "",
        "House": "NA",
        "Age": "",
        "GenderFull": ""
    }

    if not text:
        return data

    text_n = normalize_marathi(text)
    lines = [normalize_marathi(l.strip()) for l in text.split("\n") if l.strip()]

    # CardID + RegNo
    data["CardID"] = extract_card_id(text)
    data["RegNo"] = extract_reg_no(text)

    # NAME
    nm = extract_after_keyword(lines, ["मतदाराचे पूर्ण", "मतदाराचे", "नाव"])
    data["VoterName"] = clean_voter_name(nm)

    # RELATION
    rel = extract_after_keyword(
        lines,
        ["पतीचे नाव", "वडिलांचे नाव", "आईचे नाव", "तीचे नाव", "पालक"]
    )
    data["RelationName"] = clean_relative_name(rel)

    # HOUSE
    house = extract_after_keyword(lines, ["घर क्रमांक", "घर क्र", "घर"])
    data["House"] = clean_house(house)

    # AGE
    age = extract_after_keyword(lines, ["वय"])
    data["Age"] = clean_age(age)

    # GENDER
    g = extract_after_keyword(lines, ["लिंग"])
    if "पु" in g or "पुरुष" in g:
        data["GenderFull"] = "पुरुष"
    else:
        data["GenderFull"] = "महिला"

    return data


# =========================================================
# FACE CROP
# =========================================================

def crop_person_photo(card_path: Path) -> Path:
    img = Image.open(card_path)
    w, h = img.size

    left = int(w * 0.78)
    top = int(h * 0.30)
    right = int(w * 0.98)
    bottom = int(h * 0.95)

    face = img.crop((left, top, right, bottom))
    face = ImageEnhance.Contrast(face).enhance(1.2)
    face = ImageEnhance.Sharpness(face).enhance(1.3)

    out = card_path.parent / f"face_{card_path.stem}.png"
    face.save(out, "PNG")
    return out


# =========================================================
# EXCEL GENERATOR
# =========================================================

def generate_excel(all_cards):
    wb = Workbook()
    ws = wb.active
    ws.title = "Voters"

    headers = [
        "S.No.", "CardID", "RegNo",
        "मतदाराचे पूर्ण:", "Guardian",
        "घर क्रमांक :", "वय :", "लिंग :", "Face Image"
    ]

    for i, h in enumerate(headers, start=1):
        ws.cell(1, i, h).font = MARATHI_BOLD

    row = 2

    print(f"Running OCR on {len(all_cards)} cards...")

    with ProcessPoolExecutor(max_workers=os.cpu_count() or 2) as ex:
        texts = list(ex.map(ocr_card, all_cards))

    for idx, card in enumerate(all_cards):
        parsed = parse_card(texts[idx])

        ws.cell(row, 1, idx + 1)
        ws.cell(row, 2, parsed["CardID"])
        ws.cell(row, 3, parsed["RegNo"])
        ws.cell(row, 5, parsed["VoterName"])
        ws.cell(row, 6, parsed["RelationName"])
        ws.cell(row, 7, parsed["House"])
        ws.cell(row, 8, parsed["Age"])
        ws.cell(row, 9, parsed["GenderFull"])

        # FACE IMAGE
        face = crop_person_photo(card)
        if face.exists():
            img = XLImage(str(face))
            img.width, img.height = THUMB_W, THUMB_H
            ws.add_image(img, f"J{row}")
            ws.row_dimensions[row].height = THUMB_H * 0.75

        row += 1

    # column width
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        if letter == "J":
            ws.column_dimensions[letter].width = 20
            continue
        max_len = max(len(str(c.value)) for c in col if c.value)
        ws.column_dimensions[letter].width = min(max_len + 2, 50)

    wb.save(EXCEL_OUT)
    print(f"Excel saved: {EXCEL_OUT}")

    # Cleanup temp images
    cleanup_images()

# CLEANUP
def cleanup_images():
    try:
        if CARD_DIR.exists():
            time.sleep(0.1)
            shutil.rmtree(CARD_DIR)
            print(f"🧹 Cleaned up temporary folder: {CARD_DIR}")
    except Exception as e:
        print(f"⚠️ Cleanup failed: {e}")


# =========================================================
# MAIN PIPELINE
# =========================================================

if __name__ == "__main__":

    folder = Path(INPUT_IMAGE_FOLDER)
    if not folder.exists():
        raise FileNotFoundError(f"Folder not found: {folder}")

    page_images = []
    for ext in ("*.jpg", "*.jpeg", "*.png", "*.webp"):
        page_images.extend(folder.glob(ext))

    if not page_images:
        raise ValueError("No input images found.")

    cropped_pages = []
    for img in page_images:
        cropped_pages.append(crop_page(img, PAGE_DIR))

    all_cards = []
    for p in cropped_pages:
        all_cards.extend(split_page_into_cards(p, CARD_DIR))

    print(f"Total cards: {len(all_cards)}")

    generate_excel(all_cards)

    print("Done.")
