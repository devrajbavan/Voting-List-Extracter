import os
import re
import time
import uuid
import shutil
import uvicorn
import pytesseract
from io import BytesIO
from pathlib import Path
from typing import Tuple
from concurrent.futures import ProcessPoolExecutor, as_completed

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from PIL import Image, ImageEnhance

from fastapi.responses import FileResponse, JSONResponse
from fastapi import FastAPI, UploadFile, File, HTTPException, BackgroundTasks

# ---------------- CONFIG ----------------
TESSERACT_CMD = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
pytesseract.pytesseract.tesseract_cmd = TESSERACT_CMD

TESS_LANG = "mar+eng"
UPLOAD_ROOT = Path("uploads")
UPLOAD_ROOT.mkdir(exist_ok=True)

EXCEL_OUTPUT_DIR = Path("generated_excels")
EXCEL_OUTPUT_DIR.mkdir(exist_ok=True)

MARATHI_FONT = Font(name="Mangal", size=11)
MARATHI_BOLD = Font(name="Mangal", size=11, bold=True)

DEFAULT_COLS = 3
DEFAULT_ROWS = 10

FACE_LEFT_RATIO = 0.78
FACE_TOP_RATIO = 0.30
FACE_RIGHT_RATIO = 0.98
FACE_BOTTOM_RATIO = 0.85

THUMB_W, THUMB_H = 80, 90

app = FastAPI(title="API for Voter's details extraction", version="1.0")

DEV_DIGITS = "०१२३४५६७८९"
ENG_DIGITS = "0123456789"

# ---------------- OCR HELPERS ----------------

def preprocess_for_ocr(img: Image.Image) -> Image.Image:
    """Faster, lighter preprocessing."""
    img = img.convert("L")
    if img.width < 400:
        img = img.resize((img.width * 2, img.height * 2), Image.LANCZOS)
    return img

def ocr_card_text_bytes(img_bytes: bytes) -> str:
    img = Image.open(BytesIO(img_bytes))
    img = preprocess_for_ocr(img)
    txt = pytesseract.image_to_string(img, lang=TESS_LANG, config="--psm 6")
    return txt

# ---------------- DATA CLEANING ----------------

def clean_voter_name(raw: str) -> str:
    s = re.sub(r'[|¦\\\/<>]', ' ', raw)
    s = re.sub(r'\s+[A-Za-z]{1,3}\s*', ' ', s)
    s = re.sub(r'[=z&*]', '', s)
    s = re.sub(r'\s+', ' ', s).strip()
    parts = s.split()
    if len(parts) > 4:
        s = ' '.join(parts[:4])
    return s

def clean_relative_name(raw: str) -> str:
    s = re.sub(r'[|¦\\\/<>]', ' ', raw)
    s = re.sub(r'\s+[A-Za-z]{1,3}\s*', ' ', s)
    s = re.sub(r'[=z&*]', '', s)
    s = re.sub(r'\s+', ' ', s).strip()
    parts = s.split()
    if len(parts) > 3:
        s = ' '.join(parts[:3])
    return s

def clean_house(text: str) -> str:
    text = re.sub(r'[^\d०-९NAna/]', '', text)
    text = text.replace('NA', 'NA').strip()
    if not text or text == 'NA':
        return 'NA'
    m = re.search(r'[\d०-९]+', text)
    return m.group(0) if m else 'NA'

def clean_age(text: str) -> str:
    m = re.search(r'[\d०-९]+', text)
    if not m:
        return ""
    age = m.group(0)
    age = "".join(ENG_DIGITS[DEV_DIGITS.index(d)] if d in DEV_DIGITS else d for d in age)
    return age

def clean_gender(code: str) -> Tuple[str, str]:
    code = code.strip().lower()
    if 'जी' in code or 'स्त्री' in code or 'स्री' in code:
        return "स्त्री", "महिला"
    else:
        return "पु", "पुरुष"

def parse_card(text: str) -> dict:
    data = {
        "ID": "", "VoterName": "", "RelationLabel": "", "RelationName": "",
        "House": "NA", "Age": "", "GenderCode": "", "GenderFull": ""
    }
    lines = [l.strip() for l in text.split('\n') if l.strip()]

    if lines:
        m = re.search(r"([A-Z0-9]{7,})\s+(\d+/\d+/\d+)", lines[0])
        if m:
            data["ID"] = f"{m.group(1)} {m.group(2)}"
            lines = lines[1:]

    voter_pat = re.compile(r"(?:मतदाराचे?\s*(?:पूर्ण\s*))\s*[:ः]?\s*(.+)", re.I)
    for i, line in enumerate(lines):
        m = voter_pat.search(line)
        if m:
            data["VoterName"] = clean_voter_name(m.group(1))
            del lines[i]
            break
    else:
        for i, line in enumerate(lines):
            if re.search(r"(पती|वडिल|आई|सासू|पत्नी|सून)", line):
                continue
            if 'नाव' in line and not re.search(r"(घर|लिंग|वय)", line):
                parts = re.split(r'[:ः]', line, maxsplit=1)
                candidate = parts[1] if len(parts) > 1 else re.split(r'नाव', line, maxsplit=1)[-1]
                data["VoterName"] = clean_voter_name(candidate)
                del lines[i]
                break

    rel_pat_pati = re.compile(r"पतीचे\s*नाव\s*[:ः]?\s*(.+)", re.I)
    rel_pat_vadil = re.compile(r"वडिलांचे\s*नाव\s*[:ः]?\s*(.+)", re.I)
    for i, line in enumerate(lines):
        m_p = rel_pat_pati.search(line)
        if m_p:
            data["RelationLabel"] = "पतीचे नाव"
            data["RelationName"] = clean_relative_name(m_p.group(1))
            del lines[i]
            break
        m_v = rel_pat_vadil.search(line)
        if m_v:
            data["RelationLabel"] = "वडिलांचे नाव"
            data["RelationName"] = clean_relative_name(m_v.group(1))
            del lines[i]
            break

    m = re.search(r"घर\s*क्रमांक\s*[:ः]?\s*(.+)", text, re.I)
    if m:
        data["House"] = clean_house(m.group(1))

    m = re.search(r"वय\s*[:ः]?\s*([\d०-९]+)", text, re.I)
    if m:
        data["Age"] = clean_age(m.group(1))

    m = re.search(r"लिंग\s*[:ः]?\s*([^\s\r\n]+)", text, re.I)
    if m:
        data["GenderCode"], data["GenderFull"] = clean_gender(m.group(1))
    else:
        data["GenderCode"], data["GenderFull"] = "पु", "पुरुष"

    return data

# ---------------- CROP HELPERS ----------------

def crop_all_cards_from_sheet_bytes(sheet_path: Path, cols=DEFAULT_COLS, rows=DEFAULT_ROWS):
    """Crop all cards and return them as bytes instead of saving to disk."""
    big = Image.open(sheet_path)
    w, h = big.size
    cw, rh = w // cols, h // rows
    cards = []

    for r in range(rows):
        for c in range(cols):
            box = (c * cw, r * rh, (c + 1) * cw, (r + 1) * rh)
            card = big.crop(box)
            buf = BytesIO()
            card.save(buf, format="WEBP", quality=90)
            cards.append(buf.getvalue())
    return cards

def crop_person_face_bytes(card_bytes: bytes) -> BytesIO:
    """Crop face and return as BytesIO (in-memory)."""
    img = Image.open(BytesIO(card_bytes))
    w, h = img.size
    left = int(w * FACE_LEFT_RATIO)
    top = int(h * FACE_TOP_RATIO)
    right = int(w * FACE_RIGHT_RATIO)
    bottom = int(h * FACE_BOTTOM_RATIO)
    left, top = max(0, left), max(0, top)
    right, bottom = min(w, right), min(h, bottom)
    if right <= left or bottom <= top:
        return None
    face = img.crop((left, top, right, bottom))
    buf = BytesIO()
    face.save(buf, format="WEBP", quality=90)
    buf.seek(0)
    return buf

# ---------------- EXCEL GENERATION ----------------

def generate_excel_from_cards(card_bytes_list, out_xlsx: Path, start_serial: int = 9):
    """Parallel OCR + Excel writing (horizontal tabular layout)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Define header row
    headers = [
        "S.No.", "ID", "Serial", "मतदाराचे पूर्ण:", 
        "पतीचे नाव / वडिलांचे नाव", "घर क्रमांक :", 
        "वय :", "लिंग :", "Face image"
    ]

    for col_idx, header in enumerate(headers, start=1):
        ws.cell(1, col_idx, header).font = MARATHI_BOLD

    # Parallel OCR
    with ProcessPoolExecutor(max_workers=os.cpu_count()) as executor:
        futures = {executor.submit(ocr_card_text_bytes, b): b for b in card_bytes_list}
        ocr_results = []
        for future in as_completed(futures):
            try:
                txt = future.result()
                ocr_results.append(txt)
            except Exception:
                ocr_results.append("")

    serial = start_serial
    row = 2  # data starts from second row

    for idx, card_bytes in enumerate(card_bytes_list):
        txt = ocr_results[idx]
        d = parse_card(txt)

        # Face crop
        face_buf = None
        try:
            face_buf = crop_person_face_bytes(card_bytes)
        except Exception:
            face_buf = None

        # Fill row data
        ws.cell(row, 1, idx + 1).font = MARATHI_FONT               # S.No.
        ws.cell(row, 2, d.get("ID", "")).font = MARATHI_FONT       # ID
        ws.cell(row, 3, serial).font = MARATHI_FONT                # Serial
        ws.cell(row, 4, d.get("VoterName", "")).font = MARATHI_FONT
        ws.cell(row, 5, d.get("RelationName", "")).font = MARATHI_FONT
        ws.cell(row, 6, d.get("House", "")).font = MARATHI_FONT
        ws.cell(row, 7, d.get("Age", "")).font = MARATHI_FONT
        ws.cell(row, 8, d.get("GenderFull", "पुरुष")).font = MARATHI_FONT

        # Face crop and add image; immediately set row height for THIS row
        face_buf = None
        try:
            face_buf = crop_person_face_bytes(card_bytes)
        except Exception:
            face_buf = None

        if face_buf:
            try:
                img_for_excel = XLImage(face_buf)
                img_for_excel.width, img_for_excel.height = THUMB_W, THUMB_H
                cell_ref = f"I{row}"  # column I for face image
                ws.add_image(img_for_excel, cell_ref)

                # Immediately set this row's height to match image height.
                # Excel row height is in points. Approx conversion: pts = px * 0.75
                # (1 pt ≈ 1.333 px → so pts = px / 1.333 ≈ px * 0.75)
                try:
                    ws.row_dimensions[row].height = img_for_excel.height * 0.75
                except Exception:
                    # defensive: if something goes wrong, set a reasonable fallback
                    ws.row_dimensions[row].height = THUMB_H * 0.75

            except Exception as e:
                # log and continue
                print(f"⚠️ Image add failed for row {row}: {e}")

        row += 1
        serial += 1

    # Set image column width (I) once (approximate mapping)
    # openpyxl column width units != pixels; this is a pragmatic approximation.
    # Increase column width enough to fit thumbnail horizontally.
    ws.column_dimensions['I'].width = max(18, int(THUMB_W * 0.14) + 2)

    # Auto-size other columns
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        if col_letter == 'I':
            continue
        max_len = max((len(str(c.value)) for c in col if c.value), default=0)
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

    wb.save(out_xlsx)
    return out_xlsx

# ---------------- CLEANUP ----------------

def cleanup_run_dir(path: Path, delay: int = 10):
    try:
        if delay and delay > 0:
            time.sleep(delay)
        if path.exists():
            shutil.rmtree(path, ignore_errors=True)
            print(f"🧹 Cleaned up run directory: {path}")
    except Exception as e:
        print(f"⚠️ Cleanup failed for {path}: {e}")

# ---------------- API ENDPOINTS ----------------

@app.post("/process-voters/")
async def process_voters(file: UploadFile = File(...), background_tasks: BackgroundTasks = None):
    if not file.filename.lower().endswith((".jpg", ".jpeg", ".png")):
        raise HTTPException(status_code=400, detail="Only jpg/png files allowed.")

    run_id = uuid.uuid4().hex
    run_dir = UPLOAD_ROOT / run_id
    run_dir.mkdir(parents=True, exist_ok=True)

    uploaded_path = run_dir / file.filename
    try:
        with uploaded_path.open("wb") as f:
            f.write(await file.read())

        card_bytes_list = crop_all_cards_from_sheet_bytes(uploaded_path)
        out_xlsx = EXCEL_OUTPUT_DIR / f"voter_data_{run_id}.xlsx"
        generate_excel_from_cards(card_bytes_list, out_xlsx)

        if background_tasks is not None:
            background_tasks.add_task(cleanup_run_dir, run_dir, 10)
        else:
            import threading
            threading.Thread(target=cleanup_run_dir, args=(run_dir, 10), daemon=True).start()

        return JSONResponse(
            content={
                "status": "success",
                "file_name": out_xlsx.name,
                "download_url": f"/download/{run_id}"
            },
            status_code=200
        )
    except Exception as e:
        try:
            shutil.rmtree(run_dir)
        except Exception:
            pass
        return JSONResponse(content={"status": "failed", "error": str(e)}, status_code=500)

@app.get("/download/{run_id}")
async def download_file(run_id: str):
    xlsx_files = list(EXCEL_OUTPUT_DIR.glob(f"voter_data_{run_id}.xlsx"))
    if not xlsx_files:
        raise HTTPException(status_code=404, detail="Excel file not found or already deleted.")
    xlsx_path = xlsx_files[0]
    return FileResponse(
        path=xlsx_path,
        filename=xlsx_path.name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# ---------------- RUN ----------------
if __name__ == "__main__":
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)
